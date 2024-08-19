import telebot
import openpyxl
from telebot import types
import os
from datetime import datetime
import shutil

bot = telebot.TeleBot('Your token')


def load_workbook():
    global wb, sheet, order_sheet, users_chat_id
    wb = openpyxl.load_workbook(filename="database.xlsx", data_only=True)
    wb.active = 0
    sheet = wb.active
    order_sheet = wb['Заказы']
    users_chat_id = wb['Пользователи']


load_workbook()

user_catalog_position = {}
user_message_ids = {}
user_order_data = {}
admin_id = [] #your id
photo_dir = os.path.abspath('photo')
status_order=False

def user_exists(chat_id):
    for row in users_chat_id.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] == chat_id:
            return True
    return False


@bot.message_handler(commands=['start'])
def start(message):
    global status_order
    status_order = False
    markup = types.InlineKeyboardMarkup()
    button_comment = types.InlineKeyboardButton("💬Отзывы", url='https://t.me/agama_otz')
    button_support = types.InlineKeyboardButton("🔔Поддержка", url='https://t.me/dtcaree')
    button_catalog = types.InlineKeyboardButton("🛒Каталог", callback_data='show_brands')
    button_search = types.InlineKeyboardButton("🔍Поиск", callback_data='search_menu')
    markup.add(button_catalog, button_search)
    markup.add(button_comment, button_support)
    bot.send_message(message.chat.id,
                     f'Здравствуйте, {message.from_user.first_name}! \n\nБот представляет полный ассортимент кроссовок магазина Agama.\nЧтобы оформить заказ, пожалуйста, перейдите в «Каталог», выберите нужную модель кроссовок и укажите данные получателя. После этого ожидайте сообщение от менеджера для подтверждения заказа.\n\nПо заказу и вопросам, пишите по кнопке «Поддержка».',
                     reply_markup=markup)
    chat_id = message.chat.id
    if not user_exists(chat_id):
        username = message.from_user.username
        user_link = f"https://t.me/{username}" if username else "N/A"
        id_user_link = "https://web.telegram.org/k/#" + str(chat_id)
        users_chat_id.append([chat_id, user_link, id_user_link, message.chat.first_name, message.chat.last_name])
        wb.save("database.xlsx")


@bot.message_handler(commands=['get'])
def getdatabase(message):
    if message.chat.id in admin_id:
        bot.send_document(message.chat.id, open('database.xlsx', 'rb'))
    else:
        bot.send_message(message.chat.id, "У вас нет прав для использования этой команды.")


@bot.message_handler(commands=['upload'])
def upload_database(message):
    if message.chat.id in admin_id:
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        markup.add(types.KeyboardButton('Отмена'))
        msg = bot.send_message(message.chat.id, "Пожалуйста, отправьте новый файл базы данных (Excel).",
                               reply_markup=markup)
        bot.register_next_step_handler(msg, receive_new_database)
    else:
        bot.send_message(message.chat.id, "У вас нет прав для использования этой команды.")


def receive_new_database(message):
    if message.text == 'Отмена':
        bot.send_message(message.chat.id, "Операция отменена.", reply_markup=types.ReplyKeyboardRemove())
        return

    if message.chat.id in admin_id and message.document:
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)

        backup_folder = 'Backup_database'
        if not os.path.exists(backup_folder):
            os.makedirs(backup_folder)

        current_time = datetime.now().strftime('day-%Y.%m.%d_time-%H.%M.%S')
        backup_path = os.path.join(backup_folder, f"database_backup_{current_time}.xlsx")
        os.rename('database.xlsx', backup_path)

        with open('database.xlsx', 'wb') as new_file:
            new_file.write(downloaded_file)

        load_workbook()
        bot.send_message(message.chat.id, "База данных успешно обновлена и старая версия сохранена в папку бэкапов.",
                         reply_markup=types.ReplyKeyboardRemove())
    else:
        bot.send_message(message.chat.id, "Ошибка при получении файла. Попробуйте еще раз.",
                         reply_markup=types.ReplyKeyboardRemove())


@bot.message_handler(commands=['list'])
def list_backups(message):
    if message.chat.id in admin_id:
        backup_folder = 'Backup_database'
        if not os.path.exists(backup_folder):
            os.makedirs(backup_folder)

        backups = os.listdir(backup_folder)
        if backups:
            backup_list = "\n".join(backups)
            bot.send_message(message.chat.id, f"Доступные бэкапы:\n{backup_list}")
        else:
            bot.send_message(message.chat.id, "Нет доступных бэкапов.")
    else:
        bot.send_message(message.chat.id, "У вас нет прав для использования этой команды.")


@bot.message_handler(commands=['backup'])
def restore_backup(message):
    if message.chat.id in admin_id:
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        markup.add(types.KeyboardButton('Отмена'))
        msg = bot.send_message(message.chat.id, "Пожалуйста, введите имя файла бэкапа для восстановления.",
                               reply_markup=markup)
        bot.register_next_step_handler(msg, perform_restore_backup)
    else:
        bot.send_message(message.chat.id, "У вас нет прав для использования этой команды.")


def perform_restore_backup(message):
    if message.text == 'Отмена':
        bot.send_message(message.chat.id, "Операция отменена.", reply_markup=types.ReplyKeyboardRemove())
        return

    if message.chat.id in admin_id:
        backup_name = message.text
        backup_path = os.path.join('Backup_database', backup_name)

        if os.path.exists(backup_path):
            old_backup_name = f"database_old_{datetime.now().strftime('day-%Y.%m.%d_time-%H.%M.%S')}.xlsx"
            os.rename('database.xlsx', os.path.join('Backup_database', old_backup_name))
            shutil.copyfile(backup_path, 'database.xlsx')
            load_workbook()
            bot.send_message(message.chat.id, f"База данных успешно восстановлена из {backup_name}.",
                             reply_markup=types.ReplyKeyboardRemove())
        else:
            bot.send_message(message.chat.id,
                             "Указанный файл не найден. Пожалуйста, проверьте имя файла и попробуйте снова.",
                             reply_markup=types.ReplyKeyboardRemove())
    else:
        bot.send_message(message.chat.id, "У вас нет прав для использования этой команды.",
                         reply_markup=types.ReplyKeyboardRemove())

@bot.message_handler(commands=['help'])
def restore_backup(message):
    if message.chat.id in admin_id:
        bot.send_message(message.chat.id, "Список команд админа\n/start стартовое меню бота\n/send рассылка сообщения всем пользователям бота (те, кто хоть раз вводит /start)\n/get Получить базу данных \n/upload загрузить базу данных \n/list список бэкапов базы данных (если что-то сломается)\n/backup использовать бэкап из списка доступных\n/new добавить новый товар")
        bot.send_message(message.chat.id, "Есть 2 способа добавить новые товары:\n1-Через команду /new\n2-Вручную добавить фото товара в папку photo открыв на сайте хостинга папку photo и закинуть туда фотки вручную. И обновить базу данных (эксель таблицу), это тоже можно сделать либо через бота (сначала скачиваешь таблицу /get, потом добавляешь в неё товары и закидываешь обратно через команду /upload, но учти что пока ты добавляешь новые товары в эту таблицу, все заказы и новые пользователи не будут добавлятся в базу данных)")
    else:
        bot.send_message(message.chat.id, "У вас нет прав для использования этой команды.")

admin_state = {}
admin_media = {}
@bot.message_handler(commands=['new'])
def add_new_product(message):
    if message.chat.id not in admin_id:
        bot.send_message(message.chat.id, "У вас нет прав для использования этой команды.")
        return

    user_order_data[message.chat.id] = {}
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    markup.add(types.KeyboardButton('Отменить добавление товара'))
    msg = bot.send_message(message.chat.id, "Введите артикул товара.", reply_markup=markup)
    bot.register_next_step_handler(msg, process_article_step)

def process_article_step(message):
    if message.text == 'Отменить добавление товара':
        bot.send_message(message.chat.id, "Добавление товара отменено.", reply_markup=types.ReplyKeyboardRemove())
        user_order_data.pop(message.chat.id, None)
        return

    user_order_data[message.chat.id]['article'] = message.text
    msg = bot.send_message(message.chat.id, "Введите название товара.")
    bot.register_next_step_handler(msg, process_name_step)

def process_name_step(message):
    if message.text == 'Отменить добавление товара':
        bot.send_message(message.chat.id, "Добавление товара отменено.", reply_markup=types.ReplyKeyboardRemove())
        user_order_data.pop(message.chat.id, None)
        return

    user_order_data[message.chat.id]['name'] = message.text
    msg = bot.send_message(message.chat.id, "Отправьте фото товара.")
    bot.register_next_step_handler(msg, process_photo_step)

def process_photo_step(message):
    if message.text == 'Отменить добавление товара':
        bot.send_message(message.chat.id, "Добавление товара отменено.", reply_markup=types.ReplyKeyboardRemove())
        user_order_data.pop(message.chat.id, None)
        return

    if message.photo:
        file_info = bot.get_file(message.photo[-1].file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        unique_photo_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{message.photo[-1].file_id}.jpg"
        photo_path = os.path.join(photo_dir, unique_photo_name)
        with open(photo_path, 'wb') as photo_file:
            photo_file.write(downloaded_file)
        user_order_data[message.chat.id]['photo'] = unique_photo_name

        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        markup.add(types.KeyboardButton('Описания нет'))
        msg = bot.send_message(message.chat.id, "Введите описание товара (если есть).", reply_markup=markup)
        bot.register_next_step_handler(msg, process_description_step)
    else:
        msg = bot.send_message(message.chat.id, "Пожалуйста, отправьте фото товара.")
        bot.register_next_step_handler(msg, process_photo_step)

def process_description_step(message):
    if message.text == 'Отменить добавление товара':
        bot.send_message(message.chat.id, "Добавление товара отменено.", reply_markup=types.ReplyKeyboardRemove())
        user_order_data.pop(message.chat.id, None)
        return

    if message.text == 'Описания нет':
        user_order_data[message.chat.id]['description'] = ""
    else:
        user_order_data[message.chat.id]['description'] = message.text

    bot.send_message(message.chat.id, "Введите цену товара.", reply_markup=types.ReplyKeyboardRemove())
    bot.register_next_step_handler(message, process_price_step)

def process_price_step(message):
    if message.text == 'Отменить добавление товара':
        bot.send_message(message.chat.id, "Добавление товара отменено.", reply_markup=types.ReplyKeyboardRemove())
        user_order_data.pop(message.chat.id, None)
        return

    user_order_data[message.chat.id]['price'] = message.text
    msg = bot.send_message(message.chat.id, "Введите размеры товара (Например: 38 40 42 44)")
    bot.register_next_step_handler(msg, process_sizes_step)

def process_sizes_step(message):
    if message.text == 'Отменить добавление товара':
        bot.send_message(message.chat.id, "Добавление товара отменено.", reply_markup=types.ReplyKeyboardRemove())
        user_order_data.pop(message.chat.id, None)
        return

    user_order_data[message.chat.id]['sizes'] = message.text
    msg = bot.send_message(message.chat.id, "Введите ссылку на пост с этим товаром.")
    bot.register_next_step_handler(msg, process_post_link_step)

def process_post_link_step(message):
    if message.text == 'Отменить добавление товара':
        bot.send_message(message.chat.id, "Добавление товара отменено.", reply_markup=types.ReplyKeyboardRemove())
        user_order_data.pop(message.chat.id, None)
        return

    user_order_data[message.chat.id]['post_link'] = message.text
    msg = bot.send_message(message.chat.id, "Введите название бренда (пример: Nike, Adidas, Salomon, New Balance).")
    bot.register_next_step_handler(msg, process_brand_step)

def process_brand_step(message):
    if message.text == 'Отменить добавление товара':
        bot.send_message(message.chat.id, "Добавление товара отменено.", reply_markup=types.ReplyKeyboardRemove())
        user_order_data.pop(message.chat.id, None)
        return

    user_order_data[message.chat.id]['brand'] = message.text
    save_to_excel(message.chat.id)
    bot.send_message(message.chat.id, "Товар успешно добавлен.", reply_markup=types.ReplyKeyboardRemove())
    user_order_data.pop(message.chat.id, None)

def save_to_excel(chat_id):
    data = user_order_data[chat_id]
    new_row = [
        data['article'],
        data['name'],
        data['photo'],
        data['description'],
        data['price'],
        data['sizes'],
        data['post_link'],
        data['brand']
    ]
    sheet.append(new_row)
    wb.save("database.xlsx")

@bot.message_handler(commands=['send'])
def send(message):
    if message.chat.id in admin_id:
        admin_state[message.chat.id] = "awaiting_message"
        admin_media[message.chat.id] = {"photos": [], "caption": None, "text": None}

        markup = types.InlineKeyboardMarkup()
        button_cancel = types.InlineKeyboardButton("Отменить рассылку", callback_data='cancel_broadcast')
        button_sendall = types.InlineKeyboardButton("Отправить", callback_data='send_all')
        markup.add(button_cancel, button_sendall)

        bot.send_message(message.chat.id,
                         "Следующее сообщение будет отправлено всем пользователям бота. Отправьте все фото и текст в одном сообщении. После отправки всех фото и текста нажмите 'Отправить'.",
                         reply_markup=markup)
    else:
        bot.send_message(message.chat.id, "У вас нет прав для использования этой команды.")


@bot.message_handler(content_types=['text', 'photo'])
def handle_message(message):
    if message.chat.id in admin_state and admin_state[message.chat.id] == "awaiting_message":
        if message.content_type == 'photo':
            admin_media[message.chat.id]["photos"].append(types.InputMediaPhoto(message.photo[-1].file_id))
            if message.caption:
                admin_media[message.chat.id]["caption"] = message.caption
        elif message.content_type == 'text':
            admin_media[message.chat.id]["text"] = message.text


@bot.callback_query_handler(func=lambda call: call.data == 'send_all')
def send_all(call):
    if call.message.chat.id in admin_id and call.message.chat.id in admin_state:
        if not admin_media[call.message.chat.id]["photos"] and not admin_media[call.message.chat.id]["text"]:
            bot.send_message(call.message.chat.id, "Сообщения для рассылки не обнаружено.")
        else:
            chat_ids = [row[0] for row in users_chat_id.iter_rows(min_row=2, max_col=1, values_only=True)]
            if admin_media[call.message.chat.id]["photos"]:
                media_group = admin_media[call.message.chat.id]["photos"]
                if admin_media[call.message.chat.id]["caption"]:
                    media_group[0].caption = admin_media[call.message.chat.id]["caption"]
                for chat_id in chat_ids:
                    bot.send_media_group(chat_id, media_group)
            elif admin_media[call.message.chat.id]["text"]:
                for chat_id in chat_ids:
                    bot.send_message(chat_id, admin_media[call.message.chat.id]["text"])

            admin_state.pop(call.message.chat.id)
            admin_media.pop(call.message.chat.id)
            bot.send_message(call.message.chat.id, "Сообщение отправлено всем пользователям.")
    else:
        bot.send_message(call.message.chat.id, "У вас нет активной рассылки или нет прав для использования этой команды.")


@bot.callback_query_handler(func=lambda call: call.data == 'cancel_broadcast')
def cancel_broadcast(call):
    if call.message.chat.id in admin_state:
        admin_state.pop(call.message.chat.id)
        admin_media.pop(call.message.chat.id)
        bot.send_message(call.message.chat.id, "Рассылка отменена.")
    else:
        bot.send_message(call.message.chat.id, "У вас нет активной рассылки.")
@bot.callback_query_handler(func=lambda callback: callback.data == 'search_menu')
def search_menu(callback):
    markup = types.InlineKeyboardMarkup()
    button_article = types.InlineKeyboardButton("Поиск по артикулу", callback_data='search_by_article')
    button_size = types.InlineKeyboardButton("Поиск по размеру", callback_data='search_by_size')
    button_back = types.InlineKeyboardButton("🔙 Назад", callback_data='start')
    markup.add(button_article, button_size)
    markup.add(button_back)

    bot.send_message(callback.message.chat.id, "Выберите вариант поиска:", reply_markup=markup)


@bot.callback_query_handler(func=lambda callback: callback.data == 'search_by_article')
def search_by_article(callback):
    markup = types.InlineKeyboardMarkup()
    button_back = types.InlineKeyboardButton("🔙 Назад", callback_data='start')
    markup.add(button_back)

    sent_message = bot.send_message(callback.message.chat.id, "Введите артикул товара:", reply_markup=markup)
    bot.register_next_step_handler(sent_message, handle_article_search)


def handle_article_search(message):
    article = message.text
    markup = types.InlineKeyboardMarkup()
    button_back = types.InlineKeyboardButton("🔙 Назад", callback_data='start')
    markup.add(button_back)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        if row[0].value == article:
            row_data = sheet[row[0].row]
            article, name, photo_name, description, price, sizes, link, brand = [cell.value for cell in row_data]
            photo_path = os.path.join(photo_dir, photo_name)

            caption = (f"{name}\n"
                       f"{description if description else ''}\n\n"
                       f"Артикул: {article}\n\n"
                       f"Цена: {price} ₽\n\n"
                       f"Размеры в наличии (EU): {sizes}\n\n"
                       f"[Ещё фото...]({link})")

            markup = types.InlineKeyboardMarkup()
            button_order = types.InlineKeyboardButton("🛒 Оформить заказ", callback_data=f'order_{article}')
            button_back = types.InlineKeyboardButton("🔙 Назад", callback_data='start')
            markup.add(button_order, button_back)

            if os.path.exists(photo_path):
                with open(photo_path, 'rb') as photo:
                    bot.send_photo(message.chat.id, photo, caption=caption, parse_mode='Markdown', reply_markup=markup)
            else:
                bot.send_message(message.chat.id, caption, parse_mode='Markdown', reply_markup=markup)
            return

    bot.send_message(message.chat.id, "Товар с таким артикулом не найден.", reply_markup=markup)


@bot.callback_query_handler(func=lambda callback: callback.data == 'search_by_size')
def search_by_size(callback):
    sizes = set()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=6):
        if row[0].value:
            for size in row[0].value.split():
                sizes.add(size)

    markup = types.InlineKeyboardMarkup()
    for size in sizes:
        button = types.InlineKeyboardButton(size, callback_data=f'size_{size}')
        markup.add(button)
    button_back = types.InlineKeyboardButton("🔙 Назад", callback_data='start')
    markup.add(button_back)

    bot.send_message(callback.message.chat.id, "Выберите размер:", reply_markup=markup)


@bot.callback_query_handler(func=lambda callback: callback.data.startswith('size_'))
def show_catalog_by_size(callback):
    user_id = callback.from_user.id
    size = callback.data.split('_')[1]
    user_catalog_position[user_id] = {'size': size, 'index': 0}

    show_catalog_page_by_size(callback.message, user_id)


def show_catalog_page_by_size(message, user_id):
    catalog_info = user_catalog_position[user_id]
    size = catalog_info['size']
    index = catalog_info['index']

    rows = [
        (sheet[f'A{i}'].value, sheet[f'B{i}'].value, sheet[f'C{i}'].value, sheet[f'D{i}'].value, sheet[f'E{i}'].value,
         sheet[f'F{i}'].value, sheet[f'G{i}'].value)
        for i in range(2, sheet.max_row + 1)
        if size in (sheet[f'F{i}'].value or '').split()
    ]

    total_photos = len(rows)
    if total_photos == 0:
        bot.send_message(message.chat.id, f'Нет товаров для размера {size}.')
        return

    if index < 0:
        index = total_photos - 1
    elif index >= total_photos:
        index = 0

    user_catalog_position[user_id]['index'] = index

    article, name, photo_name, description, price, sizes, link = rows[index]
    photo_path = os.path.join(photo_dir, photo_name)

    caption = (f"[[{index + 1} из {total_photos}]]\n\n"
               f"{name}\n"
               f"{description if description else ''}\n\n"
               f"Артикул: {article}\n\n"
               f"Цена: {price} ₽\n\n"
               f"Размеры в наличии (EU): {sizes}\n\n"
               f"[Ещё фото...]({link})")

    markup = types.InlineKeyboardMarkup()
    button_prev = types.InlineKeyboardButton("⬅️ Назад", callback_data=f'catalog_size_prev')
    button_next = types.InlineKeyboardButton("Вперед ➡️", callback_data=f'catalog_size_next')
    button_order = types.InlineKeyboardButton("🛒 Оформить заказ", callback_data=f'order_{article}')
    button_back = types.InlineKeyboardButton("🔙 Назад", callback_data='start')
    markup.add(button_prev, button_next)
    markup.add(button_order)
    markup.add(button_back)

    if user_id in user_message_ids:
        for msg_id in user_message_ids[user_id]:
            try:
                bot.delete_message(message.chat.id, msg_id)
            except telebot.apihelper.ApiException as e:
                if e.error_code != 400:
                    print(f"Не удалось удалить сообщение: {e}")

    user_message_ids[user_id] = []

    if os.path.exists(photo_path):
        with open(photo_path, 'rb') as photo:
            sent_message = bot.send_photo(message.chat.id, photo, caption=caption, parse_mode='Markdown', reply_markup=markup)
            user_message_ids[user_id].append(sent_message.message_id)
    else:
        sent_message = bot.send_message(message.chat.id, caption, parse_mode='Markdown', reply_markup=markup)
        user_message_ids[user_id].append(sent_message.message_id)


@bot.callback_query_handler(func=lambda callback: callback.data == 'catalog_size_prev')
def catalog_size_prev(callback):
    user_id = callback.from_user.id
    user_catalog_position[user_id]['index'] -= 1
    show_catalog_page_by_size(callback.message, user_id)


@bot.callback_query_handler(func=lambda callback: callback.data == 'catalog_size_next')
def catalog_size_next(callback):
    user_id = callback.from_user.id
    user_catalog_position[user_id]['index'] += 1
    show_catalog_page_by_size(callback.message, user_id)


@bot.callback_query_handler(func=lambda callback: callback.data == 'show_brands')
def show_brands(callback):
    brands = set()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=8, max_col=8):
        brand = row[0].value
        if brand:
            brands.add(brand)

    markup = types.InlineKeyboardMarkup()
    for brand in brands:
        button = types.InlineKeyboardButton(brand, callback_data=f'brand_{brand}')
        markup.add(button)
    button_back = types.InlineKeyboardButton("🔙 Назад", callback_data='start')
    markup.add(button_back)

    bot.send_message(callback.message.chat.id, "Выберите бренд:", reply_markup=markup)


@bot.callback_query_handler(func=lambda callback: callback.data.startswith('brand_'))
def show_catalog(callback):
    user_id = callback.from_user.id
    brand = callback.data.split('_')[1]
    user_catalog_position[user_id] = {'brand': brand, 'index': 0}

    show_catalog_page(callback.message, user_id)


def show_catalog_page(message, user_id):
    catalog_info = user_catalog_position[user_id]
    brand = catalog_info['brand']
    index = catalog_info['index']

    rows = [
        (sheet[f'A{i}'].value, sheet[f'B{i}'].value, sheet[f'C{i}'].value, sheet[f'D{i}'].value, sheet[f'E{i}'].value,
         sheet[f'F{i}'].value, sheet[f'G{i}'].value)
        for i in range(2, sheet.max_row + 1)
        if sheet[f'H{i}'].value == brand
    ]

    total_photos = len(rows)
    if total_photos == 0:
        bot.send_message(message.chat.id, f'Нет товаров для бренда {brand}.')
        return

    if index < 0:
        index = total_photos - 1
    elif index >= total_photos:
        index = 0

    user_catalog_position[user_id]['index'] = index

    article, name, photo_name, description, price, sizes, link = rows[index]

    photo_path = os.path.join(photo_dir, photo_name)

    caption = (f"[[{index + 1} из {total_photos}]]\n\n"
               f"{name}\n"
               f"{description if description else ''}\n\n"
               f"Артикул: {article}\n\n"
               f"Цена: {price} ₽\n\n"
               f"Размеры в наличии (EU): {sizes}\n\n"
               f"[Ещё фото...]({link})")

    markup = types.InlineKeyboardMarkup()
    button_prev = types.InlineKeyboardButton("⬅️ Назад", callback_data='catalog_prev')
    button_next = types.InlineKeyboardButton("Вперед ➡️", callback_data='catalog_next')
    button_order = types.InlineKeyboardButton("🛒 Оформить заказ", callback_data=f'order_{article}')
    button_back = types.InlineKeyboardButton("🔙 Назад", callback_data='start')
    markup.add(button_prev, button_next)
    markup.add(button_order)
    markup.add(button_back)

    if user_id in user_message_ids:
        for msg_id in user_message_ids[user_id]:
            try:
                bot.delete_message(message.chat.id, msg_id)
            except telebot.apihelper.ApiException as e:
                if e.error_code != 400:
                    print(f"Не удалось удалить сообщение: {e}")

    user_message_ids[user_id] = []

    if os.path.exists(photo_path):
        with open(photo_path, 'rb') as photo:
            sent_message = bot.send_photo(message.chat.id, photo, caption=caption, parse_mode='Markdown', reply_markup=markup)
            user_message_ids[user_id].append(sent_message.message_id)
    else:
        sent_message = bot.send_message(message.chat.id, caption, parse_mode='Markdown', reply_markup=markup)
        user_message_ids[user_id].append(sent_message.message_id)


@bot.callback_query_handler(func=lambda callback: callback.data == 'catalog_prev')
def catalog_prev(callback):
    user_id = callback.from_user.id
    user_catalog_position[user_id]['index'] -= 1
    show_catalog_page(callback.message, user_id)


@bot.callback_query_handler(func=lambda callback: callback.data == 'catalog_next')
def catalog_next(callback):
    user_id = callback.from_user.id
    user_catalog_position[user_id]['index'] += 1
    show_catalog_page(callback.message, user_id)


@bot.callback_query_handler(func=lambda callback: callback.data.startswith('order_'))
def callback_order(callback):
    article = callback.data.split('_')[1]
    user_order_data[callback.from_user.id] = {'article': article}

    markup = types.InlineKeyboardMarkup()
    row = next(row for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6) if row[0].value == article)
    sizes = row[5].value.split()
    for size in sizes:
        markup.add(types.InlineKeyboardButton(size, callback_data=f'select_size_{size}'))
    markup.add(types.InlineKeyboardButton("🔙 Назад", callback_data='start'))

    bot.send_message(callback.message.chat.id, "Выберите размер:", reply_markup=markup)


@bot.callback_query_handler(func=lambda callback: callback.data.startswith('select_size_'))
def select_size(callback):
    global status_order
    status_order=True
    size = callback.data.split('_')[2]
    user_order_data[callback.from_user.id]['size'] = size
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Сбросить", callback_data='start'))

    bot.send_message(callback.message.chat.id, "Для оформления заказа от вас потребуется данные получателя. Пожалуйста, отправьте ФИО в ответном сообщении.", reply_markup=markup)
    bot.register_next_step_handler(callback.message, handle_name)


def handle_name(message):
    global status_order
    if (status_order):
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Сбросить", callback_data='start'))
        user_order_data[message.from_user.id]['name'] = message.text
        bot.send_message(message.chat.id, "Выберите метод доставки:", reply_markup=delivery_method_markup())


def delivery_method_markup():
    global status_order
    if (status_order):
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("СДЭК", callback_data='delivery_method_cdek'), types.InlineKeyboardButton("Почта России", callback_data='delivery_method_russian_post'))
        markup.add(types.InlineKeyboardButton("Сбросить", callback_data='start'))
        return markup


@bot.callback_query_handler(func=lambda callback: callback.data.startswith('delivery_method_'))
def select_delivery_method(callback):
    global status_order
    if (status_order):
        if callback.data=="delivery_method_cdek":
            method="CДЭК"
        else:
            method = "Почта России"
        user_order_data[callback.from_user.id]['delivery_method'] = method
        bot.send_message(callback.message.chat.id, "Введите город получателя.")
        bot.register_next_step_handler(callback.message, handle_city)


def handle_city(message):
    global status_order
    if (status_order):
        user_order_data[message.from_user.id]['city'] = message.text
        bot.send_message(message.chat.id, "Введите адрес пункта выдачи.")
        bot.register_next_step_handler(message, handle_address)


def handle_address(message):
    global status_order
    if (status_order):
        user_order_data[message.from_user.id]['address'] = message.text
        bot.send_message(message.chat.id, "Введите индекс.")
        bot.register_next_step_handler(message, handle_postcode)


def handle_postcode(message):
    global status_order
    if (status_order):
        user_order_data[message.from_user.id]['postcode'] = message.text
        bot.send_message(message.chat.id, "Укажите телефон получателя (в формате +79997778899).")
        bot.register_next_step_handler(message, handle_phone)


def handle_phone(message):
    global status_order
    if (status_order):
        user_order_data[message.from_user.id]['phone'] = message.text
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Да", callback_data='cod_yes'))
        markup.add(types.InlineKeyboardButton("Нет", callback_data='cod_no'))

        bot.send_message(message.chat.id, "Посылка с наложенным платежом", reply_markup=markup)


@bot.callback_query_handler(func=lambda callback: callback.data.startswith('cod_'))
def handle_cod(callback):
    global status_order
    if (status_order):
        cod = callback.data.split('_')[1] == 'yes'
        user_order_data[callback.from_user.id]['cod'] = cod
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        markup.add(types.KeyboardButton('Пропустить'))
        if not(cod):
            bot.send_message(callback.message.chat.id, "Пришлите скриншот квитанции/чека с оплатой заказа @dtcaree (https://t.me/dtcaree).\n\nРеквизиты для оплаты заказа:\n\nСбербанк: 2202206780046103\nТинькофф: 2200700766083710\n\nНа имя: Данил Алексеевич А. (Akulov Danil)")
        bot.send_message(callback.message.chat.id, "Комментарий к заказу (если есть).", reply_markup=markup)
        bot.register_next_step_handler(callback.message, handle_comment)


def handle_comment(message):
    global status_order
    if (status_order):
        user_order_data[message.from_user.id]['comment'] = message.text
        confirm_order(message)




def confirm_order(message):
    global status_order
    if (status_order):
        chat_id = message.chat.id
        username = message.from_user.username
        user_link = f"https://t.me/{username}" if username else "N/A"
        id_user_link = "https://web.telegram.org/k/#" + str(chat_id)

        user_order_data[message.from_user.id]['user_link']=user_link
        user_order_data[message.from_user.id]['id_link']=id_user_link
        user_order_data[message.from_user.id]['first_name'] = message.chat.first_name
        user_order_data[message.from_user.id]['last_name'] = message.chat.last_name

        try:
            order_data = user_order_data[chat_id]
        except KeyError:
            return
        article = order_data['article']
        size = order_data['size']
        name = order_data['name']
        delivery_method = order_data['delivery_method']
        city = order_data.get('city', '')
        address = order_data['address']
        postcode = order_data['postcode']
        phone = order_data['phone']
        cod = 'Да' if order_data['cod'] else 'Нет'
        comment = order_data.get('comment', '')

        confirmation_text = (f"Проверьте правильность введенных данных.\n\n"
                             f"Получатель: {name}\n"
                             f"Артикул: {article}\n"
                             f"Размер: {size}\n"
                             f"Метод доставки: {delivery_method}\n"
                             f"Город: {city}\n"
                             f"Адрес: {address}\n"
                             f"Индекс: {postcode}\n"
                             f"Телефон: {phone}\n"
                             f"Наложенный платеж: {cod}\n"
                             f"Комментарий: {comment}")

        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Подтвердить", callback_data='confirm_order'))
        markup.add(types.InlineKeyboardButton("Сбросить", callback_data='start'))

        bot.send_message(message.chat.id, confirmation_text, reply_markup=markup)

def generate_order_number():
    order_numbers = [cell.value for cell in order_sheet['O'][1:] if cell.value is not None]
    return max(order_numbers) + 1 if order_numbers else 1

@bot.callback_query_handler(func=lambda callback: callback.data == 'confirm_order')
def complete_order(callback):
    global status_order
    if status_order:
        order_data = user_order_data.pop(callback.from_user.id, None)
        if order_data['user_link'] == "N/A":
            bot.send_message(callback.message.chat.id, "Пожалуйста, напишите свой номер заказа @dtcaree (https://t.me/dtcaree), чтобы мы смогли с вами связаться. Спасибо!")
        if order_data:
            order_number = generate_order_number()
            new_row = [order_data['name'], "", order_data['article'], order_data['size'], order_data['delivery_method'], order_data.get('city', ''),
                       order_data['address'], order_data['postcode'], order_data['phone'], "", 'Да' if order_data['cod'] else 'Нет',
                       order_data.get('comment', ''), order_data['user_link'],order_data['id_link'], order_number, order_data['first_name'],order_data['last_name']]
            bot.send_message(admin_id[1], f"Новый заказ!\n"
                                          f"Номер заказа: {order_number}\n"
                                          f"ФИО:{order_data['name']}\n"
                                          f"Артикул:{order_data['article']}\n"
                                          f"Размер: {order_data['size']}\n"
                                          f"Метод доставки: {order_data['delivery_method']}\n"
                                          f"Город: {order_data.get('city', '')}\n"
                                          f"Адрес: {order_data['address']}\n"
                                          f"Индекс: {order_data['postcode']}\n"
                                          f"Телефон: {order_data['phone']}\n"
                                          f"Наложенный платеж: {'Да' if order_data['cod'] else 'Нет'}\n"
                                          f"Комментарий: {order_data.get('comment', '')}\n"
                                          f"Ссылка на аккаунт: {order_data['user_link']}\n"
                                          f"Ник в тг: {order_data['first_name']} {order_data['last_name']}")
            order_sheet.append(new_row)
            wb.save("database.xlsx")

            bot.send_message(callback.message.chat.id, f"Ваш заказ оформлен! Ожидайте подтверждения заказа от менеджера. Номер вашего заказа: {order_number}", reply_markup=types.ReplyKeyboardRemove())
        else:
            bot.send_message(callback.message.chat.id, "Произошла ошибка. Попробуйте оформить заказ заново.",reply_markup=types.ReplyKeyboardRemove())
        back_to_start(callback)

@bot.callback_query_handler(func=lambda callback: callback.data == 'start')
def back_to_start(callback):
    start(callback.message)

bot.polling(none_stop=True)